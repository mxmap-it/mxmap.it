#!/usr/bin/env python3
"""Build a single, citizen-friendly dataset of the mxmap.it scan results.

Produces 3 representations of the same flat per-entity table, all under
the `dist/` directory at the repo root (deployed via GitHub Pages):

  dist/mxmap_it_dataset.csv   — UTF-8 with BOM (Excel/LibreOffice will
                                 auto-detect the encoding); semicolon-
                                 separated for Italian-locale Excel.
  dist/mxmap_it_dataset.json  — same rows + a metadata header object
                                 for API/programmatic consumers.
  dist/mxmap_it_dataset.xlsx  — Excel workbook with 4 sheets:
                                  1. Enti (flat, filterable)
                                  2. Per categoria IPA
                                  3. Per cluster cittadino
                                  4. Funnel trasparenza
                                 Generated only if openpyxl is installed.

Each row joins:
  - data/municipalities_it.json   (IndicePA seed — codice_ipa, ISTAT,
                                   denominazione, region, district)
  - data.json                     (DNS classification per entity)
  - data/it_citizen_clusters.json (cluster mapping per IPA category)

Schema (per-entity flat row):
  codice_ipa, codice_categoria, categoria_label, cluster_key,
  cluster_label, denominazione, codice_istat, codice_comune_istat,
  regione, provincia, sito_istituzionale, domain_used, domain_source,
  has_mx, provider_raw, provider_display, sovereignty_bucket,
  mx_records, mx_countries, gateway, dkim_tenant, spf_includes,
  classification_reason, ipa_url

Run:
  uv run python3 scripts/build_public_dataset.py

Auto-updated on every pipeline run via:
  scripts/run_it_pipeline.sh           (step 14b)
  scripts/server_autorun_full_pipeline.sh (after build_frontend)
"""
from __future__ import annotations

import csv
import io
import json
import sys
import time
from pathlib import Path
from typing import Any

ROOT = Path(__file__).resolve().parent.parent
DATA = ROOT / "data"
DIST = ROOT / "dist"
DIST.mkdir(exist_ok=True)

CSV_OUT  = DIST / "mxmap_it_dataset.csv"
JSON_OUT = DIST / "mxmap_it_dataset.json"
XLSX_OUT = DIST / "mxmap_it_dataset.xlsx"

# Citizen-facing display name (matches index.html / build_frontend)
PROVIDER_DISPLAY = {
    "microsoft": "Microsoft 365", "google": "Google Workspace", "aws": "AWS",
    "aruba": "Provider Italiano", "register-it": "Provider Italiano",
    "seeweb": "Provider Italiano", "infocert": "Provider Italiano",
    "namirial": "Provider Italiano", "local-isp": "Provider Italiano",
    "telia": "Provider Italiano", "tet": "Provider Italiano",
    "zone": "Provider Italiano", "elkdata": "Provider Italiano",
    "pa-contractor-private": "Provider Italiano",
    "regional-public": "Cloud Italiano",
    "independent": "Infrastruttura autonoma",
    "provincial-shared": "Mail provinciale condivisa",
    "zoho": "Zoho", "yandex": "Yandex",
    "unknown": "Sconosciuto",
}

# Sovereignty bucket = high-level digital-sovereignty categorisation,
# the same used in stats panel.
def sovereignty_of(provider_display: str) -> str:
    if provider_display in {"Microsoft 365", "Google Workspace", "AWS"}:
        return "USA (CLOUD Act)"
    if provider_display == "Cloud Italiano":
        return "Italia — Cloud sovrano"
    if provider_display in {"Provider Italiano", "Mail provinciale condivisa"}:
        return "Italia — Provider commerciali"
    if provider_display == "Infrastruttura autonoma":
        return "Italia — Infrastruttura autonoma"
    if provider_display in {"Zoho", "Yandex"}:
        return "Altri provider esteri"
    return "Sconosciuto"


# Friendly Italian labels for IPA category codes (citizen-facing).
CATEGORY_LABEL = {
    "L4":  "Regione / Provincia autonoma",
    "L5":  "Provincia",
    "L45": "Città metropolitana",
    "L6":  "Comune",
    "L33": "Scuola statale",
    "L17": "Università pubblica",
    "L43": "Istituzione AFAM",
    "L15": "Ente diritto allo studio universitario",
    "L28": "Consorzio interuniversitario di ricerca",
    "L7":  "ASL",
    "L8":  "Azienda ospedaliera / Policlinico",
    "L22": "IRCCS",
    "C12": "Istituto Zooprofilattico",
    "C1":  "Ministero",
    "C2":  "Avvocatura / Organo costituzionale",
    "C5":  "Autorità amministrativa indipendente",
    "C10": "Agenzia fiscale",
    "C11": "Forza di polizia",
    "L46": "Azienda dello Stato a ordinamento autonomo",
    "C14": "Ordine professionale",
    "L35": "Camera di Commercio",
    "C13": "Automobile Club federato ACI",
    "L34": "Azienda Pubblica di Servizi alla Persona (ex-IPAB)",
    "L39": "Azienda/Consorzio Edilizia Residenziale Pubblica",
    "L38": "Parco / Autorità di bacino",
    "L40": "Consorzio sviluppo industriale",
    "L42": "Ente regolazione servizi idrici/rifiuti",
    "L44": "Comunità isolane / parchi regionali",
    "L31": "Teatro stabile",
    "L16": "Fondazione Lirico-Sinfonica",
    "C7":  "Ente pubblico servizi assistenziali/culturali",
    "C8":  "Ente di ricerca pubblico (CNR, INFN, INAF, ENEA, …)",
    "L13": "Agenzia regionale sviluppo agricolo",
    "L21": "Agenzia regionale erogazioni in agricoltura",
    "L11": "Autorità di Sistema Portuale",
    "L18": "Unione di Comuni",
    "L36": "Comunità montana / Consorzio BIM",
    "L12": "Consorzio tra Amministrazioni Locali",
    "L24": "Altro ente locale",
    "L1":  "Amministrazione separata Beni Civici",
    "L47": "Commissario straordinario governativo",
    "L20": "Comunità isolana",
    "L2":  "Agenzia regionale formazione/ricerca/ambiente",
    "L19": "Agenzia regionale del lavoro",
    "L10": "Agenzia/Ente turismo regionale",
    "L37": "Gestore di pubblici servizi",
    "S01": "Società in conto economico consolidato",
    "S01G": "Società consolidata (sotto-categoria)",
    "C3":  "Ente pubblico non economico",
    "SA":  "Stazione appaltante",
    "SAG": "Stazione appaltante aggregatrice",
}


def main() -> int:
    seed_path = DATA / "municipalities_it.json"
    data_path = ROOT / "data.json"
    clusters_path = DATA / "it_citizen_clusters.json"
    if not seed_path.exists() or not data_path.exists():
        print("FATAL: seed or data.json missing — run pipeline first")
        return 1

    seed = json.loads(seed_path.read_text(encoding="utf-8"))
    data = json.loads(data_path.read_text(encoding="utf-8"))
    muns = data.get("municipalities", {})

    clusters = {}
    cat_to_cluster: dict[str, tuple[str, str]] = {}  # cat -> (cluster_key, cluster_label)
    if clusters_path.exists():
        cdoc = json.loads(clusters_path.read_text(encoding="utf-8"))
        clusters = cdoc.get("clusters", {})
        for ckey, cinfo in clusters.items():
            label = cinfo.get("label_it") or ckey
            for cat in cinfo.get("categories", []):
                cat_to_cluster[cat] = (ckey, label)

    # ISTAT province name lookup (3-digit prefix -> name)
    # Reuse what fetch_indicepa exports inside the seed (district field already
    # holds the OSM province name for L6 entities).

    rows: list[dict[str, Any]] = []
    seed_by_id = {e.get("id"): e for e in seed if e.get("id")}
    for bid, m in muns.items():
        if not bid or not bid.startswith(("IT-",)):
            continue
        sd = seed_by_id.get(bid, {})
        cat = sd.get("ipa_codice_categoria") or ""
        ckey, clabel = cat_to_cluster.get(cat, ("", ""))
        provider_raw = m.get("provider") or "unknown"
        provider_disp = PROVIDER_DISPLAY.get(provider_raw, provider_raw)
        codice_ipa = sd.get("ipa_codice_ipa") or ""
        ipa_url = f"https://indicepa.gov.it/ipa-portale/visualizza-scheda/{codice_ipa}" if codice_ipa else ""

        # Spaccato gateway / dkim / spf — solo i campi citizen-rilevanti
        gateway = m.get("gateway") or ""
        dkim_tenant = ""
        if isinstance(m.get("dkim"), dict):
            for sel, target in m["dkim"].items():
                if "onmicrosoft" in (target or "").lower():
                    dkim_tenant = target
                    break
        spf_includes = ""
        spf = m.get("spf") or ""
        if spf:
            # rough extraction of include: tokens
            includes = [t.split(":", 1)[1] for t in spf.split() if t.startswith("include:")]
            spf_includes = ";".join(includes[:8])

        mx_hosts = m.get("mx") or []
        mx_countries = m.get("mx_countries") or []

        rows.append({
            "codice_ipa":           codice_ipa,
            "codice_categoria":     cat,
            "categoria_label":      CATEGORY_LABEL.get(cat, cat),
            "cluster_key":          ckey,
            "cluster_label":        clabel,
            "denominazione":        m.get("name") or sd.get("name") or "",
            "codice_istat":         sd.get("ipa_codice_istat") or "",
            "codice_comune_istat":  sd.get("ipa_codice_comune_istat") or "",
            "regione":              sd.get("region") or m.get("canton") or "",
            "provincia":            sd.get("district") or m.get("district") or "",
            "sito_istituzionale":   sd.get("domain") or "",
            "domain_used":          m.get("domain") or "",
            "domain_source":        sd.get("domain_source") or "",
            "has_mx":               "true" if mx_hosts else "false",
            "provider_raw":         provider_raw,
            "provider_display":     provider_disp,
            "sovereignty_bucket":   sovereignty_of(provider_disp),
            "mx_records":           ";".join(mx_hosts[:10]),
            "mx_countries":         ";".join(mx_countries),
            "gateway":              gateway,
            "dkim_tenant":          dkim_tenant,
            "spf_includes":         spf_includes,
            "classification_reason": m.get("reason") or "",
            "ipa_url":              ipa_url,
        })

    # Sort: provider_display asc, then denominazione for stable diff-friendly order
    rows.sort(key=lambda r: (r["provider_display"], r["denominazione"].lower()))

    fields = list(rows[0].keys()) if rows else []
    generated = time.strftime("%Y-%m-%dT%H:%M:%SZ", time.gmtime())

    # ---- CSV (UTF-8 BOM, semicolon for IT-locale Excel) ----
    with open(CSV_OUT, "wb") as f:
        f.write(b"\xef\xbb\xbf")  # UTF-8 BOM
        sio = io.StringIO()
        w = csv.DictWriter(sio, fieldnames=fields, delimiter=";",
                           quoting=csv.QUOTE_MINIMAL,
                           extrasaction="ignore", lineterminator="\n")
        w.writeheader()
        for r in rows:
            w.writerow(r)
        f.write(sio.getvalue().encode("utf-8"))
    print(f"  CSV  : {CSV_OUT}  ({CSV_OUT.stat().st_size:,} bytes, {len(rows)} rows)")

    # ---- JSON (with metadata header) ----
    payload = {
        "_meta": {
            "name": "mxmap.it — Italian PA Email Sovereignty Dataset",
            "generated": generated,
            "source": "https://github.com/fpietrosanti/mxmap.it",
            "license": "ODbL-1.0 + CC-BY-4.0 (data); MIT (code)",
            "rows": len(rows),
            "fields": fields,
            "indicepa_source": "https://indicepa.gov.it/ipa-dati/dataset/enti",
        },
        "rows": rows,
    }
    JSON_OUT.write_text(json.dumps(payload, ensure_ascii=False, indent=2),
                        encoding="utf-8")
    print(f"  JSON : {JSON_OUT}  ({JSON_OUT.stat().st_size:,} bytes)")

    # ---- XLSX (Excel multi-sheet) — optional ----
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        wb = Workbook()
        ws = wb.active
        ws.title = "Enti"
        ws.append(fields)
        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="1A1A2E")
            cell.alignment = Alignment(horizontal="left")
        for r in rows:
            ws.append([r.get(k, "") for k in fields])
        # Auto-width approximation
        for col_idx, name in enumerate(fields, 1):
            ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = \
                min(40, max(12, len(name) + 2))
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions

        # Per categoria sheet
        from collections import Counter, defaultdict
        ws2 = wb.create_sheet("Per categoria IPA")
        cat_count = Counter(r["codice_categoria"] for r in rows)
        cat_provs: dict[str, Counter] = defaultdict(Counter)
        for r in rows:
            cat_provs[r["codice_categoria"]][r["provider_display"]] += 1
        ws2.append(["codice_categoria", "categoria_label", "cluster_label", "n_enti",
                    "dominante", "dominante_pct", "usa_pct"])
        for cat, n in cat_count.most_common():
            label = CATEGORY_LABEL.get(cat, cat)
            cluster = cat_to_cluster.get(cat, ("", ""))[1]
            provs = cat_provs[cat]
            top = provs.most_common(1)[0] if provs else ("?", 0)
            usa = sum(provs.get(p, 0) for p in ("Microsoft 365", "Google Workspace", "AWS"))
            ws2.append([cat, label, cluster, n, top[0],
                        round(top[1]/n*100, 1) if n else 0,
                        round(usa/n*100, 1) if n else 0])
        for cell in ws2[1]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill("solid", fgColor="DDDDDD")
        ws2.column_dimensions["B"].width = 40
        ws2.column_dimensions["C"].width = 28
        ws2.freeze_panes = "A2"

        # Per cluster sheet
        ws3 = wb.create_sheet("Per cluster")
        clu_count: Counter[str] = Counter()
        clu_label: dict[str, str] = {}
        clu_provs: dict[str, Counter] = defaultdict(Counter)
        for r in rows:
            ck = r["cluster_key"] or "(none)"
            clu_count[ck] += 1
            clu_label[ck] = r["cluster_label"] or "(non assegnato)"
            clu_provs[ck][r["provider_display"]] += 1
        ws3.append(["cluster_key", "cluster_label", "n_enti", "dominante",
                    "dominante_pct", "usa_pct", "cloud_italiano_pct"])
        for ck, n in clu_count.most_common():
            provs = clu_provs[ck]
            top = provs.most_common(1)[0] if provs else ("?", 0)
            usa = sum(provs.get(p, 0) for p in ("Microsoft 365", "Google Workspace", "AWS"))
            cit = provs.get("Cloud Italiano", 0)
            ws3.append([ck, clu_label[ck], n, top[0],
                        round(top[1]/n*100, 1) if n else 0,
                        round(usa/n*100, 1) if n else 0,
                        round(cit/n*100, 2) if n else 0])
        for cell in ws3[1]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill("solid", fgColor="DDDDDD")
        ws3.column_dimensions["B"].width = 40
        ws3.freeze_panes = "A2"

        # Funnel sheet
        ws4 = wb.create_sheet("Funnel trasparenza")
        funnel_path = DATA / "it_pipeline_funnel.json"
        if funnel_path.exists():
            funnel = json.loads(funnel_path.read_text(encoding="utf-8"))
            ws4.append(["", funnel.get("summary_sentence", "")])
            ws4.append([])
            ws4.append(["Stage", "Conteggio", "Percentuale"])
            for s in funnel.get("stages", []):
                ws4.append([s.get("label", ""), s.get("count", 0), f"{s.get('pct', 0):.2f}%"])
            for cell in ws4[3]:
                cell.font = Font(bold=True)
                cell.fill = PatternFill("solid", fgColor="DDDDDD")
            ws4.column_dimensions["A"].width = 60
        else:
            ws4["A1"] = "data/it_pipeline_funnel.json non presente."

        wb.save(XLSX_OUT)
        print(f"  XLSX : {XLSX_OUT}  ({XLSX_OUT.stat().st_size:,} bytes)")
    except ImportError:
        print(f"  XLSX : skipped (openpyxl not installed). pip install openpyxl to enable.")

    print(f"\nDataset built. {len(rows)} rows. Generated: {generated}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
